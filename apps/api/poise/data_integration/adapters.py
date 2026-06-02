"""SourceAdapter 抽象基类 + 内置实现 + 注册表。

设计目的（系统设计 §6.2）：
把"接客户/银行/ERP"这件事抽成一个稳定接口 —— 客户/IT 只需写一个 Adapter
子类，把异构数据源（SAP / Oracle / 用友 / 金蝶 / 银企直连 / Excel）
映射成 Canonical Model。核心引擎不变。

提供：
- 抽象基类 `SourceAdapter`
- 简单适配器 `CsvDirectoryAdapter`（目录下找标准 7 表 CSV，等价 demo_company）
- Excel 适配器 `ExcelWorkbookAdapter`（一个 .xlsx 7 个 sheet）
- 适配器注册表 `ADAPTER_REGISTRY`：name → 类，便于 API/CLI 按名调度

不在此处实现的（生产里由 IT 各家自写）：
- ERPSapAdapter / ERPOracleAdapter / ERPYongyouAdapter / ERPKingdeeAdapter
- BankApiAdapter（招行/工行/建行 等银企直连）

参考下面 ExampleStub 为模板。
"""

from __future__ import annotations

from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, ClassVar


# ============================================================
# Canonical 数据契约（dict 形态，对齐 CSV 字段）
# ============================================================
#
# 每个适配器把外部数据映射为下列 dict 列表，再交给 importers 写库。
# 这与 CSV 导入器读到的 dict 结构 1:1 对齐。


CanonicalRow = dict[str, Any]


@dataclass
class CanonicalDataset:
    """适配器输出的完整数据契约。"""

    entities: list[CanonicalRow] = field(default_factory=list)
    accounts: list[CanonicalRow] = field(default_factory=list)
    balances: list[CanonicalRow] = field(default_factory=list)
    cashflows: list[CanonicalRow] = field(default_factory=list)
    instruments: list[CanonicalRow] = field(default_factory=list)
    credit_lines: list[CanonicalRow] = field(default_factory=list)
    reserve_rules: list[CanonicalRow] = field(default_factory=list)

    def total_rows(self) -> int:
        return sum(
            len(getattr(self, t))
            for t in ("entities", "accounts", "balances", "cashflows",
                      "instruments", "credit_lines", "reserve_rules")
        )

    def summary(self) -> dict[str, int]:
        return {
            t: len(getattr(self, t))
            for t in ("entities", "accounts", "balances", "cashflows",
                      "instruments", "credit_lines", "reserve_rules")
        }


# ============================================================
# 抽象基类
# ============================================================


class SourceAdapter(ABC):
    """所有数据源适配器的契约。"""

    name: ClassVar[str] = "abstract"

    @abstractmethod
    def fetch(self, **kwargs: Any) -> CanonicalDataset:
        """从源系统拉取数据，返回 Canonical 形态。

        kwargs 视具体实现而定：
          - CSV：path
          - Excel：path
          - SAP：connection_str / company_code / date_range
          - 银行：account_no / from_date / to_date
        """
        raise NotImplementedError


# ============================================================
# CsvDirectoryAdapter ：目录下找 7 个标准 CSV
# ============================================================


class CsvDirectoryAdapter(SourceAdapter):
    """从目录读取标准 7 表 CSV。文件名固定：

    entities.csv / accounts.csv / balances.csv / cashflows.csv
    instruments.csv / credit_lines.csv / reserve_rules.csv

    与 demo_company seed 完全等价；客户只要按这套字段名出 CSV 就能接入。
    """

    name = "csv_directory"

    def fetch(self, *, path: str | Path) -> CanonicalDataset:
        import csv

        base = Path(path)
        if not base.is_dir():
            raise FileNotFoundError(f"路径不是目录：{base}")

        ds = CanonicalDataset()
        for table in (
            "entities", "accounts", "balances", "cashflows",
            "instruments", "credit_lines", "reserve_rules",
        ):
            f = base / f"{table}.csv"
            if not f.exists():
                continue
            with f.open("r", encoding="utf-8") as fh:
                rows = list(csv.DictReader(fh))
            setattr(ds, table, rows)
        return ds


# ============================================================
# ExcelWorkbookAdapter ：一个 .xlsx，7 个 sheet
# ============================================================


class ExcelWorkbookAdapter(SourceAdapter):
    """从一个 Excel 工作簿读取数据，每张表对应一个 sheet（同名）。

    要求：openpyxl 包（pip install openpyxl）。
    若客户用 Excel 而非 CSV（常见），此 adapter 是最快接入路径。
    """

    name = "excel_workbook"

    def fetch(self, *, path: str | Path) -> CanonicalDataset:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise RuntimeError(
                "需要 openpyxl 才能读 Excel：pip install openpyxl"
            ) from e

        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        ds = CanonicalDataset()
        for table in (
            "entities", "accounts", "balances", "cashflows",
            "instruments", "credit_lines", "reserve_rules",
        ):
            if table not in wb.sheetnames:
                continue
            ws = wb[table]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c).strip() if c is not None else "" for c in rows[0]]
            data: list[CanonicalRow] = []
            for r in rows[1:]:
                # 全空行跳过
                if all(c is None or c == "" for c in r):
                    continue
                row_dict = {h: ("" if v is None else
                                (v.isoformat() if isinstance(v, date) else str(v)))
                            for h, v in zip(header, r) if h}
                data.append(row_dict)
            setattr(ds, table, data)
        return ds


# ============================================================
# ExampleStub ：生产 ERP 适配器的模板
# ============================================================


class ExampleERPStub(SourceAdapter):
    """生产 ERP 适配器的样例骨架（不可直接用）。

    复制此类，重命名（如 ErpSapAdapter），填入：
    1. __init__：接收 connection_str / company_code 等连接参数
    2. fetch()：连源系统，按字段 mapping 转 CanonicalDataset

    示例字段 mapping（SAP）：
        SAP BKPF.BUKRS → entity.code
        SAP BSEG.BSCHL='40'（借）/'50'（贷）→ direction
        SAP BSEG.DMBTR → amount
        SAP BSEG.BUDAT → expected_date
    """

    name = "example_erp_stub"

    def __init__(self, connection_str: str = "", company_code: str = "DEMO") -> None:
        self.connection_str = connection_str
        self.company_code = company_code

    def fetch(self, **kwargs: Any) -> CanonicalDataset:
        # 生产实现：用 SAP RFC / sqlalchemy / requests 等连接源系统
        raise NotImplementedError(
            "ExampleERPStub 只是模板；请复制此类并实现 fetch() 方法。"
            "字段 mapping 参考 doc/数据契约_CSV.md"
        )


# ============================================================
# 增量 upsert helper —— 适配器层共用
# ============================================================


def merge_canonical(base: CanonicalDataset, incoming: CanonicalDataset) -> CanonicalDataset:
    """把 incoming 合并入 base（按自然键 upsert）。

    自然键：
      entities         (code,)
      accounts         (entity_code, code)
      balances         (entity_code, account_code, as_of_date)
      cashflows        (entity_code, expected_date, direction, category, counterparty, amount)
        ※ cashflow 没强自然键，用复合最稳；客户也可改加 external_id 字段
      instruments      (entity_code, code)
      credit_lines     (entity_code, code)
      reserve_rules    (entity_code,)
    """

    def upsert(target: list[CanonicalRow], incoming_rows: list[CanonicalRow],
               key_fn: Callable[[CanonicalRow], tuple]) -> list[CanonicalRow]:
        idx = {key_fn(r): i for i, r in enumerate(target)}
        for row in incoming_rows:
            k = key_fn(row)
            if k in idx:
                target[idx[k]] = row   # 替换
            else:
                target.append(row)
                idx[k] = len(target) - 1
        return target

    out = CanonicalDataset(
        entities=list(base.entities),
        accounts=list(base.accounts),
        balances=list(base.balances),
        cashflows=list(base.cashflows),
        instruments=list(base.instruments),
        credit_lines=list(base.credit_lines),
        reserve_rules=list(base.reserve_rules),
    )
    upsert(out.entities, incoming.entities, lambda r: (r.get("code"),))
    upsert(out.accounts, incoming.accounts,
           lambda r: (r.get("entity_code"), r.get("code")))
    upsert(out.balances, incoming.balances,
           lambda r: (r.get("entity_code"), r.get("account_code"), r.get("as_of_date")))
    upsert(out.cashflows, incoming.cashflows,
           lambda r: (r.get("entity_code"), r.get("expected_date"), r.get("direction"),
                      r.get("category"), r.get("counterparty"), r.get("amount")))
    upsert(out.instruments, incoming.instruments,
           lambda r: (r.get("entity_code"), r.get("code")))
    upsert(out.credit_lines, incoming.credit_lines,
           lambda r: (r.get("entity_code"), r.get("code")))
    upsert(out.reserve_rules, incoming.reserve_rules,
           lambda r: (r.get("entity_code"),))
    return out


# ============================================================
# 适配器注册表
# ============================================================


ADAPTER_REGISTRY: dict[str, type[SourceAdapter]] = {
    CsvDirectoryAdapter.name: CsvDirectoryAdapter,
    ExcelWorkbookAdapter.name: ExcelWorkbookAdapter,
    # 真实 ERP 适配器在客户/IT 实现后用 register() 注册
}


def register(adapter_cls: type[SourceAdapter]) -> None:
    """允许第三方实现注册自家 adapter。"""
    ADAPTER_REGISTRY[adapter_cls.name] = adapter_cls


def get_adapter(name: str) -> SourceAdapter:
    if name not in ADAPTER_REGISTRY:
        raise KeyError(f"未注册的 adapter：{name}；已注册：{list(ADAPTER_REGISTRY)}")
    return ADAPTER_REGISTRY[name]()
